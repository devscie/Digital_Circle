from openpyxl import Workbook

filename = "panilha_nova.xlsx"

workbook = Workbook()
sheet = workbook.active

sheet["A1"] = 1
sheet["A2"] = 2
sheet["A3"] = 3
sheet["A4"] = 4
sheet["A5"] = 5

# Save the spreadsheet
#workbook.save(filename=filename)

#from openpyxl import load_workbook
#filename = "panilha_nova.xlsx"

# Start by opening the spreadsheet and selecting the main sheet
#workbook = load_workbook(filename=filename)
#sheet = workbook.active

# Write what you want into a specific cell
sheet["A6"] = "=SUM(A1:A5)"

# Save the spreadsheet
workbook.save(filename=filename)

#workbook = load_workbook(filename=filename)
print(workbook.sheetnames)

sheet2 = workbook.create_sheet()

print(sheet2)
print(sheet2.title)

print(workbook.sheetnames)

#nova_sheet = workbook["Sheet1"]
sheet2.title = "Aba 2"

print(workbook.sheetnames)

sheet2["C3 "] = "Teste 123"

workbook.save(filename=filename)